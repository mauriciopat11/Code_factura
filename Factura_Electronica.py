from __future__ import print_function
import pickle
import os.path
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.http import MediaIoBaseDownload
import io
import re
from os import listdir
from os.path import isfile, join
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.styles.colors import WHITE
import datetime

# If modifying these scopes, delete the file token.pickle.
SCOPES = ['https://www.googleapis.com/auth/drive']

def main():
    """Shows basic usage of the Drive v3 API.
    Prints the names and ids of the first 10 files the user has access to.
    """
    creds = None
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)

    service = build('drive', 'v3', credentials=creds)

    # Call the Drive v3 API
    results = service.files().list(q="'1bBWxuYTZbMg5UFK8La5-cUrcScKM2lre' in parents",
        pageSize=1000, fields="nextPageToken, files(id, name)").execute()
#    results = service.files().list(q="'1MQO1Z0i-txOmbteB-VNyw9C4xs2xON5m' in parents",
#        pageSize=1000, fields="nextPageToken, files(id, name)").execute()

    items = results.get('files', [])

    # Crear Diccionario Facturas fileId : Name
    factura_dic = {}
    if not items:
        print('No files found.')
    else:
        print('Files:')
        for item in items:
            print(u'{0} ({1})'.format(item['name'], item['id']))
            factura_dic.update({item['id'] : item ['name']})
    return service, factura_dic

def download_file(service, factura_dic):
    for file_id, name in factura_dic.items():
        request = service.files().get_media(fileId=file_id)
        fh = io.FileIO("/Users/mpatinob/Dropbox/Personal/Negocios/Lili Pink/Facturas_Electronica/" + name, 'wb')
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
            print ("Download %d%%." % int(status.progress() * 100))

def Parse_facturas():
    fact_proc = []
    path = '/Users/mpatinob/Dropbox/Personal/Negocios/Lili Pink/Facturas_Electronica'
    files = [f for f in listdir(path) if isfile(join(path, f))]
    for f in files:
        subtotal = 0
        descuento = 0
        if '.xml' in f:
            print (f)
            factura =  open('/Users/mpatinob/Dropbox/Personal/Negocios/Lili Pink/Facturas_Electronica/' + str(f))
            factura_parse = factura.read()
            Factura_id = re.findall(r"<cbc:ParentDocumentID>(\w+)<\/cbc:ParentDocumentID>", factura_parse)
            if len(Factura_id) == 0:
                Factura_id = ["N/A"]
            Tienda = re.findall(r"<cbc:CityName>(\w+|\w+ \w+)<\/cbc:CityName>", factura_parse)
            if len(Tienda) == 0:
                Tienda = ["N/A"]
            Issue_Date = re.findall(r"<cbc:IssueDate>(\d+-\d+-\d+)<\/cbc:IssueDate>", factura_parse)
            if len(Issue_Date) == 0:
                Issue_Date = ["N/A", "N/A"]
            Credito = re.findall(r"<cbc:Description>(CREDITOS LILI PINK)</cbc:Description>", factura_parse)
            if len(Credito) == 0:
                Credito.append('CONTADO')
            Valor_total = re.findall(r'<CustomField Name="PayAmount" Value="(\d+\.\d+)" \/>' , factura_parse)
            if len(Valor_total) == 0:
                Valor_total = ['0']
            iva = re.findall(r'<CustomField Name="TotalImpuestos" Value="(\d+\.\d+)" \/>', factura_parse)
            if len(iva) == 0:
                iva = ['0']
            if 'CREDITO' in Credito[0]:
                interes = re.findall(r'<cbc:PriceAmount currencyID="COP">(\d+\.\d+)<\/cbc:PriceAmount>', factura_parse)
            else:
                interes = ['0' , '0']
            subtotal = round((float(Valor_total[0]) - float(iva[0]))/0.63 , 2)
            descuento = round(subtotal * 0.37 , 2)
            fact_proc.append([Factura_id[0],Tienda[0],Issue_Date[1], str(subtotal), str(descuento), iva[0], Valor_total[0], interes[1], Credito[0]])
    return fact_proc

def Reporte(fact_proc):
    excel_salida = openpyxl.Workbook()

    blackFill = PatternFill(start_color='00000000', end_color='00000000', fill_type='solid')
    lightgreyFill = PatternFill(start_color='E0E0E0E0', end_color='E0E0E0E0', fill_type='solid')
    redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    excel_salida.create_sheet(index=0, title='Facturas')
    hoja_salida = excel_salida.worksheets[0]

    hoja_salida['A1'] = 'No Factura'
    hoja_salida['B1'] = 'Tienda'
    hoja_salida['C1'] = 'Fecha de Expedicion'
    hoja_salida['D1'] = 'Interes'
    hoja_salida['E1'] = 'Subtotal'
    hoja_salida['F1'] = 'Descuento'
    hoja_salida['G1'] = 'IVA'
    hoja_salida['H1'] = 'Total'
    hoja_salida['I1'] = 'Tipo Fact'

    hoja_salida['A1'].fill = blackFill
    hoja_salida['A1'].font = Font(color=WHITE, bold=True, size=18)
    hoja_salida['B1'].fill = blackFill
    hoja_salida['B1'].font = Font(color=WHITE, bold=True, size=18)
    hoja_salida['C1'].fill = blackFill
    hoja_salida['C1'].font = Font(color=WHITE, bold=True, size=18)
    hoja_salida['D1'].fill = blackFill
    hoja_salida['D1'].font = Font(color=WHITE, bold=True, size=18)
    hoja_salida['E1'].fill = blackFill
    hoja_salida['E1'].font = Font(color=WHITE, bold=True, size=18)
    hoja_salida['F1'].fill = blackFill
    hoja_salida['F1'].font = Font(color=WHITE, bold=True, size=18)
    hoja_salida['G1'].fill = blackFill
    hoja_salida['G1'].font = Font(color=WHITE, bold=True, size=18)
    hoja_salida['H1'].fill = blackFill
    hoja_salida['H1'].font = Font(color=WHITE, bold=True, size=18)
    hoja_salida['I1'].fill = blackFill
    hoja_salida['I1'].font = Font(color=WHITE, bold=True, size=18)

    dim = {'A': 20, 'B': 20, 'C': 30 , 'D': 20 , 'E': 20, 'F': 20, 'G': 20, 'H' : 20, 'I' : 20}
    for k, v in dim.items():
        hoja_salida.column_dimensions[k].width = v

    row = 2

    for fact in fact_proc:
        hoja_salida['A' + str(row)] = fact[0]
        hoja_salida['B' + str(row)] = fact[1]
        hoja_salida['C' + str(row)] = fact[2]
        hoja_salida['D' + str(row)] = fact[7].replace(".", ",")
        hoja_salida['E' + str(row)] = fact[3].replace(".", ",")
        hoja_salida['F' + str(row)] = fact[4].replace(".", ",")
        hoja_salida['G' + str(row)] = fact[5].replace(".", ",")
        hoja_salida['H' + str(row)] = fact[6].replace(".", ",")
        hoja_salida['I' + str(row)] = fact[8]
        row += 1

    fecha = datetime.date.today()
    excel_salida.save("/Users/mpatinob/Dropbox/Personal/Negocios/Lili Pink/Facturas_Electronica/Facturas/Facturas_" + str(fecha) + ".xlsx")

if __name__ == '__main__':
    service, factura_dic = main()
    download_file(service, factura_dic)
    fact_proc = Parse_facturas()
    Reporte(fact_proc)
